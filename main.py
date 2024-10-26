import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import Protection
from openpyxl.utils import column_index_from_string
import json
import os


class ExcelProtectionApp:
    def __init__(self, master):
        self.master = master
        self.master.title("Excel Protection")
        self.master.resizable(True, False)  # Make the window only horizontally resizable

        self.file_path = tk.StringVar()
        self.password = tk.StringVar()
        self.row_nums = tk.StringVar()
        self.col_letters = tk.StringVar()
        self.show_password = tk.BooleanVar()
        self.protect_formulas = tk.BooleanVar(value=True)  # Default to checked

        self.password_entry = None  # Define here to initialize in create_widgets

        self.load_data()  # Load data on startup

        self.create_widgets()
        self.master.grid_rowconfigure(0, weight=1)
        self.master.grid_columnconfigure(0, weight=1)

        self.master.protocol("WM_DELETE_WINDOW", self.on_closing)  # Save data on close

    def create_widgets(self):
        outer_frame = tk.Frame(self.master, padx=10, pady=10)
        outer_frame.grid(row=0, column=0, sticky="nsew")

        # Configure resizing for the outer frame
        outer_frame.grid_rowconfigure(0, weight=1)
        outer_frame.grid_columnconfigure(1, weight=1)

        tk.Label(outer_frame, text="File Path:").grid(row=0, column=0, sticky="w")
        file_frame = tk.Frame(outer_frame)
        file_frame.grid(row=0, column=1, columnspan=3, sticky="ew")
        tk.Entry(file_frame, textvariable=self.file_path, width=50).pack(side=tk.LEFT, expand=True, fill=tk.X)
        tk.Button(file_frame, text="...", command=self.browse_file, width=2, height=1, padx=2).pack(side=tk.RIGHT,
                                                                                                    padx=1, pady=1)

        tk.Label(outer_frame, text="Password:").grid(row=1, column=0, sticky="w")
        password_frame = tk.Frame(outer_frame)
        password_frame.grid(row=1, column=1, columnspan=3, sticky="ew")
        self.password_entry = tk.Entry(password_frame, textvariable=self.password, show='*')
        self.password_entry.pack(side=tk.LEFT, expand=True, fill=tk.X)
        tk.Checkbutton(password_frame, text="Show", variable=self.show_password, command=self.toggle_password).pack(
            side=tk.RIGHT)

        tk.Label(outer_frame, text="Row Numbers (comma-separated):").grid(row=2, column=0, sticky="w")
        tk.Entry(outer_frame, textvariable=self.row_nums).grid(row=2, column=1, columnspan=3, sticky="ew")

        tk.Label(outer_frame, text="Column Letters (comma-separated):").grid(row=3, column=0, sticky="w")
        tk.Entry(outer_frame, textvariable=self.col_letters).grid(row=3, column=1, columnspan=3, sticky="ew")

        tk.Checkbutton(outer_frame, text="Protect cells with formulas", variable=self.protect_formulas).grid(row=4,
                                                                                                             column=0,
                                                                                                             columnspan=4,
                                                                                                             sticky="w")

        button_frame = tk.Frame(outer_frame)
        button_frame.grid(row=5, column=0, columnspan=4, sticky="e")
        tk.Button(button_frame, text="Protect", command=self.protect_file).pack(side=tk.LEFT, padx=5, pady=10)
        tk.Button(button_frame, text="Unprotect", command=self.unprotect_file).pack(side=tk.LEFT, padx=5, pady=10)

    def toggle_password(self):
        if self.show_password.get():
            self.password_entry.config(show='')
        else:
            self.password_entry.config(show='*')

    def browse_file(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        self.file_path.set(filename)

    def protect_file(self):
        try:
            self.protect_cells_all_sheets(
                self.file_path.get(),
                row_nums=[int(n) for n in self.row_nums.get().split(',')] if self.row_nums.get() else None,
                col_letters=self.col_letters.get().split(',') if self.col_letters.get() else None,
                password=self.password.get(),
                protect_formulas=self.protect_formulas.get()
            )
            messagebox.showinfo("Success", "File protected successfully!")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def unprotect_file(self):
        try:
            self.unprotect_cells_all_sheets(self.file_path.get(), password=self.password.get())
            messagebox.showinfo("Success", "File unprotected successfully!")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def save_data(self):
        data = {
            "file_path": self.file_path.get(),
            "password": self.password.get(),
            "row_nums": self.row_nums.get(),
            "col_letters": self.col_letters.get(),
            "show_password": self.show_password.get(),
            "protect_formulas": self.protect_formulas.get(),
            "window_size": (self.master.winfo_width(), self.master.winfo_height())
        }
        with open("user_data.json", "w") as f:
            json.dump(data, f)

    def load_data(self):
        if os.path.exists("user_data.json"):
            with open("user_data.json", "r") as f:
                data = json.load(f)
                self.file_path.set(data.get("file_path", ""))
                self.password.set(data.get("password", ""))
                self.row_nums.set(data.get("row_nums", ""))
                self.col_letters.set(data.get("col_letters", ""))
                self.show_password.set(data.get("show_password", False))
                self.protect_formulas.set(data.get("protect_formulas", True))
                window_size = data.get("window_size", (0, 0))
                if window_size[0] > 0:  # previous window_size available
                    self.master.geometry(f"{window_size[0]}x{window_size[1]}")

    def on_closing(self):
        self.save_data()
        self.master.destroy()

    @staticmethod
    def protect_cells_all_sheets(file_path, row_nums=None, col_letters=None, password=None, protect_formulas=False):
        workbook = load_workbook(file_path)
        for sheet in workbook:
            for row in sheet.iter_rows():
                for cell in row:
                    cell.protection = Protection(locked=False)
            if row_nums:
                for row_num in row_nums:
                    for cell in sheet[row_num]:
                        cell.protection = Protection(locked=True)
            if col_letters:
                col_nums = [column_index_from_string(letter) for letter in col_letters]
                for col_num in col_nums:
                    for row in sheet.iter_rows(min_col=col_num, max_col=col_num):
                        for cell in row:
                            cell.protection = Protection(locked=True)
            if protect_formulas:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.data_type == 'f':
                            cell.protection = Protection(locked=True)
            if password:
                sheet.protection.sheet = True
                sheet.protection.password = password
        workbook.save(file_path)

    @staticmethod
    def unprotect_cells_all_sheets(file_path, password=None):
        workbook = load_workbook(file_path)
        for sheet in workbook:
            for row in sheet.iter_rows():
                for cell in row:
                    cell.protection = Protection(locked=False)
            sheet.protection.sheet = False
            if password:
                sheet.protection.password = ''
        workbook.save(file_path)


if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelProtectionApp(root)
    root.mainloop()
